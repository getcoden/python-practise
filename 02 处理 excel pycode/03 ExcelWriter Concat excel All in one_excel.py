# -*- coding:utf-8 -*-


# pandas 将多个 dataframe 以多个 sheet 的形式保存到一个 excel 文件中

import glob
import os
import time
from openpyxl import load_workbook
import pandas as pd


writer = pd.ExcelWriter('test.xlsx')
data1.to_excel(writer, sheet_name='sheet1')
data2.to_excel(writer, sheet_name='sheet2')
writer.save()

# 上面的方法会将原来的 excel 文件覆盖掉，假如想要对已经存在的 excel 文件进行修改，可以使用开源工具包（anaconda 已附带）openpyxl
# 这种方法只能是创建一个空的不含任何数据的工作薄，如果含有旧的工作表也将会被覆盖


writer = pd.ExcelWriter('test.xlsx', engin='openpyxl')
book = load_workbook(writer.path)
writer.book = book
dataframe.to_excel(excel_writer=writer, sheet_name="info5")
writer.save()
writer.close()

# 如果想用的旧的工作薄作为文件，而且又不想里的工作数据被覆盖，只能用下面这种方法


# 根据时间创建工作表名子

path = r"d:\test1\excel"
if not os.path.exists(path):
    os.mkdir(path)

# 工作表名字
date_rand = time.strftime("%Y-%m-%d", time.localtime())
file_name = path+'\%s.xlsx' % date_rand
writer = pd.ExcelWriter(file_name, engine='openpyxl')
# 加上这一句是为了不覆盖掉原xlsx文件里的工作表,要加入 engine='openpyxl' 引擎
book = load_workbook(writer.path)
# 加上这一句是为了不覆盖掉原xlsx文件里的工作表，同时也 import 入 from openpylx import load_workbook
writer.book = book


df1 = pd.DataFrame([[15, 100, 167], [18, 120, 178]], columns=[
                   'id', 'age', 'subject'])  # columns和inex参数可要可不要
df2 = pd.DataFrame([[24, 134, 175], [35, 140, 180]])
# 写入缓存
df1.to_excel(writer, 'df11', index=False)
df2.to_excel(writer, 'df22', index=False)

# # 将缓存写入工作表
writer.save()


# 这两个例子都是利用新创建的工作薄来保存工作表而不会出现覆盖工作表情况

all_data = pd.DataFrame()
path = r"C:\Users\win\Desktop\2\*.xlsx"
for f in glob.glob(path):
    df = pd.read_excel(f, sheet_name=None)
    cdf = pd.concat(df.values())
    all_data = all_data.append(cdf)
print(all_data)
df.to_excel(r"C:\Users\win\Desktop\2\cocant.xlsx")


# 合并多 excel 文件到一个文件的多个 sheet
# 利用 Pandas 将一个文件夹下的多个 excel 文件合并成一个 excel 文件，原来的每个 excel 文件作为新文件的一个 sheet


# 需要创建ExcelWriter对象，要不sheet会覆盖，只留最后一个

writer = pd.ExcelWriter(r"C:\Users\win\Desktop\2\test_new.xlsx")
flist = glob.glob(r"C:\Users\win\Desktop\2\*.xlsx")
for i in flist:
    df = pd.read_excel(i)
    # 提取文件名
    fname = i.split("\\")[-1]
    fname = fname.split(".")[0]
    print(fname)
    df.to_excel(writer, sheet_name=fname)
writer.save()

"""
  import pandas as pd
  import os
  from openpyxl import load_workbook  # 必须导入的模块

  path = r"D:\test1\file_folder"
  listing = os.listdir(path)
  data_list = [f for f in listing if f[-5:] == '.xlsx']  # 为了排除不是.xlsx的文件，避免程序报错

  with pd.ExcelWriter(r"D:\test1\result.xlsx", engine='openpyxl') as writer:  # 这是一种高效的写法
      for i in data_list:
          name = i.split('.')[0]
          df1 = pd.read_excel(os.path.join(path, i))
          df1.to_excel(writer, sheet_name=name, index=False)  # 程序的逻辑

  # 另一种写法

  writer = pd.ExcelWriter(r'result.xlsx')

  for i in os.listdir(r'./file_folder'):
      name = i.split('.')[0]
      df1 = pd.read_excel(os.path.join(r'./file_folder', i))
      df1.to_excel(writer, sheet_name=name, index=False)
  writer.save()  # 这2个操作不能忘记，否则文件只在缓存里，看不到直接的操作结果
  writer.close()


  # 获取工作薄所有工作表列表
  # 方法一
  # df = pd.read_excel(path, sheet_name=None)
  # for i in df.keys():
  #     print(i)


  # 方法二:
  # df = pd.read_excel(path, sheet_name=None)
  # sheet_name_list = list(df)
  # print(sheet_name_list)


    """

# w.save(time.strftime('%Y_%m_%d_%H_%M_%S', time.localtime(
#     time.time())) + " excelName.xls")  # 根据时间来保存文件名


# 读取excel 某个sheet下的数据,将其转化为列表数据，方便操作


def get_excel_sheet_data(filename, sheetname):
    try:
        df = pd.read_excel(filename, sheet_name=sheetname)
        dataList = df.to_dict(orient='records')  # 转换为列表
        return dataList
    except:
        print('未能打开此文件，请确认文件名及路径是否正确')
        return []

# 读取excel 所有sheet表名，并将其转化为列表返回


def get_excel_sheet_list(filename):
    try:
        df = pd.read_excel(filename, sheet_name=None)
        return df.keys()
    except Exception as e:
        print(e)
        return []


if __name__ == '__main__':
    fileName = '数据.xlsx'
    sheetList = get_excel_sheet_list(fileName)  # 获取sheet列表
    print(sheetList)
    writer = pd.ExcelWriter("新数据.xlsx")
    for i in sheetList:  # 遍历每个sheet名称
        sheetData = get_excel_sheet_data(fileName, i)   # 读取每个sheet下的数据
        for j in sheetData:  # 取出每一项的数据
            # do something
            # 依次根据sheet名称对“新数据.xlsx”此文件多次写入Sheet
        pd_look = pd.DataFrame(sheetData)
        pd_look.to_excel(writer, sheet_name=i)
        # 最后保存写入，并释放
    writer.save()
    writer.close()

# i） 合并同一文件夹下的所有工作簿


def concat_xlsx(dir_name, result_filename='result'):
    '''合并同一文件夹下的所有excel工作簿'''
    dfs = []
    for filename in os.listdir(dir_name):
        if os.path.splitext(filename)[1] == '.xlsx':
            full_path = os.path.join(dir_name, filename)
            df = pd.read_excel(full_path)
            dfs.append(df)

    result = pd.concat(dfs)
    result_path = os.path.join(dir_name, result_filename + '.xlsx')
    result.to_excel(result_path, index=False, freeze_panes=(1, 0))

# iii） 合并一个工作簿中的多个工作表


def concat_worksheets(path, result_path):
    '''合并同一工作簿中的所有工作表'''
    dfs = pd.read_excel(path, sheet_name=None).values()
    result = pd.concat(dfs)
    result.to_excel(result_path, index=0, freeze_panes=(1, 0))
