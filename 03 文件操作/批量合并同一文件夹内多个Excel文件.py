import time
import pandas as pd
import os
import glob
from openpyxl import load_workbook

all_data = pd.DataFrame()
path = r"C:\Users\win\Desktop\2\*.xlsx"
for f in glob.glob(path):
    df = pd.read_excel(f, sheet_name=None)
    cdf = pd.concat(df.values())
    all_data = all_data.append(cdf)
print(all_data)
df.to_excel(r"C:\Users\win\Desktop\2\cocant.xlsx")


# 合并多 excel 文件到一个文件的多个 sheet
# 利用 Pandas 将一个文件夹下的多个 excel 文件合并成一个 excel 文件，原来的每个 excel 文件作为新文件的一个 sheet


# 需要创建ExcelWriter对象，要不sheet会覆盖，只留最后一个

writer = pd.ExcelWriter(r"C:\Users\win\Desktop\2\test_new.xlsx")
flist = glob.glob(r"C:\Users\win\Desktop\2\*.xlsx")
for i in flist:
    df = pd.read_excel(i)
    # 提取文件名
    fname = i.split("\\")[-1]
    fname = fname.split(".")[0]
    print(fname)
    df.to_excel(writer, sheet_name=fname)
writer.save()


workbook_url = 'https://github.com/chris1610/pbpython/raw/master/data/2018_Sales_Total_Tabs.xlsx'
all_dfs = pd.read_excel(workbook_url, sheet_name=None)
all_dfs.to_excel(r'd:\test1\all_dfs_concat.xlsx')
print(all_dfs.keys())
df = pd.concat(all_dfs, ignore_index=True)
print(df)


# 根据时间创建工作表名子

path = r"d:\test1\excel"
if not os.path.exists(path):
    os.mkdir(path)

# 工作表名字
date_rand = time.strftime("%Y-%m-%d", time.localtime())
file_name = path+'\%s.xlsx' % date_rand
writer = pd.ExcelWriter(file_name, engine='openpyxl')
# 加上这一句是为了不覆盖掉原xlsx文件里的工作表,要加入 engine='openpyxl' 引擎
book = load_workbook(writer.path)
# 加上这一句是为了不覆盖掉原xlsx文件里的工作表，同时也 import 入 from openpylx import load_workbook
writer.book = book


df1 = pd.DataFrame([[15, 100, 167], [18, 120, 178]], columns=[
                   'id', 'age', 'subject'])  # columns和inex参数可要可不要
df2 = pd.DataFrame([[24, 134, 175], [35, 140, 180]])
# 写入缓存
df1.to_excel(writer, 'df11', index=False)
df2.to_excel(writer, 'df22', index=False)

# # 将缓存写入工作表
writer.save()


# 将多个表格合并在一个sheet中

writer2 = pd.ExcelWriter(r"d:\test1\excel\ceshi.xlsx")

data1 = [1, 2, 3, 4]
data2 = [5, 6, 7, 8]

df1 = pd.DataFrame({'name': data1})
df2 = pd.DataFrame({'name': data2})

df1.to_excel(writer2)
df2.to_excel(writer2, startcol=5, startrow=10)  # 表示第10行 第5列

writer2.save()
