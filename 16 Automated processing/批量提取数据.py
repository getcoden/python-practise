from openpyxl import load_workbook, Workbook

# 数据所在的文件夹目录
path = 'D:/test1'

# 打开电商婴儿数据工作簿
workbook = load_workbook(
    path + '/' + '(sample)sam_tianchi_mum_baby_trade_history.xlsx')
# 打开工作表
sheet = workbook.active
buy_mount = sheet['F']
row_lst = []
for cell in buy_mount:
    if isinstance(cell.value, int) and cell.value > 50:
        # print(cell.row)
        row_lst.append(cell.row)
new_workbook = Workbook()
new_sheet = new_workbook.active

# 创建和 电商婴儿数据 一样的表头（第一行）
header = sheet[1]
header_lst = []
for cell in header:
    header_lst.append(cell.value)
new_sheet.append(header_lst)

# 从旧表中根据行号提取符合条件的行，并遍历单元格获取值，以列表形式写入新表
for row in row_lst:
    data_lst = []
    for cell in sheet[row]:
        data_lst.append(cell.value)
    new_sheet.append(data_lst)

# 最后切记保存
new_workbook.save(path + '/' + '符合筛选条件的新表.xlsx')
